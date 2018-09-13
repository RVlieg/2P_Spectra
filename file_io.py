# -*- coding: utf-8 -*-
"""
Created on Tue Jun 26 14:37:47 2018

@author: vlieg
"""
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog


#%% Change Extension of a String
def change_extension(path_file, extension):
    path_old = path_file    
    extension_index = str.find(path_old, '.')    
    if str.find(extension,'.') != 0:
        extension = '.' + extension        
    path_new = path_old[0:extension_index] + extension

    return path_new

#%% Get filepath from a pop-up window
def get_path():
    root = tk.Tk()
    root.withdraw()

    filename = filedialog.askopenfilename()
    filename = filename.replace('/','\\')

    return filename 


#%%
def read_logfile(path_logfile):
    path_logfile = change_extension(path_logfile,'.log')
    f = open(path_logfile, 'r')
    log = f.readlines()[:]
    f.close()
    
    for n, line in enumerate(log):
        log[n] = line.rstrip()
        
    return log

#%% Read 16-bit .bin file as 2D image 
def read_bin(path_binfile,slice_nr):
    path_binfile = change_extension(path_binfile,'.bin')
    
    path_logfile = change_extension(path_binfile,'.log')
    data_log = read_logfile(path_logfile)
    
    # Get 2D dimensions image from log file 
    ypix = np.uint(str.split(data_log[8],'=')[1])
    xpix = np.uint(str.split(data_log[9],'=')[1])
    pix_slice = ypix*xpix
    
    # Open file and read data of 1 slice
    f = open(path_binfile, 'rb')
    
    if slice_nr!=0:
       f.seek(slice_nr*pix_slice*2,0)
       im_slice = np.fromfile(f, count = pix_slice, dtype = '>i2')
    else:
       im_slice = np.fromfile(f, count = pix_slice, dtype = '>i2')  
        
    f.close()
    im_slice = np.reshape(im_slice,[ypix,xpix],'F')

    return im_slice


#%% Read Specified Stack from .bin file 
def read_stack(filepath,stack_nr):
    logfile= read_logfile(filepath)
    
    ypix   = np.uint(str.split(logfile[8],'=')[1])
    xpix   = np.uint(str.split(logfile[9],'=')[1])
    zsteps = logfile[18]
    zsteps = str.split(zsteps,"=")[1]
    zsteps = np.uint(str.split(zsteps,",")[0])
    stack=np.zeros([xpix,ypix,zsteps], dtype = np.uint16)
        
    for slice_nr in range(stack_nr*zsteps,stack_nr*zsteps + zsteps):
        stack[:,:,slice_nr-stack_nr*zsteps]=read_bin(filepath,slice_nr)
    
    return stack 

#%%
def read_bin_all(filepath):
    logfile= read_logfile(filepath)
    
    ypix   = np.uint(str.split(logfile[8],'=')[1])
    xpix   = np.uint(str.split(logfile[9],'=')[1])
    nframes = np.uint(str.split(logfile[7],'=')[1])
    
    images =np.zeros([xpix,ypix,nframes], dtype = np.uint16)
        
    for slice_nr in range(0,nframes):
        images[:,:,slice_nr]=read_bin(filepath,slice_nr)
    
    return images 

#%% Write List formatted Data to an .xlsx file 
def write_xlsx_list(file_name,data_list,column_headers,column_indices):
    
    file_name = change_extension(file_name,'.xlsx')
   
    data=pd.DataFrame(data_list)
    num_stacks = len(data)
    num_param = np.shape(data[0][0])[1]
    
    
    try: 
        book = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
    except FileNotFoundError:
        print('No .xlsx file found')        
        writer = pd.ExcelWriter(file_name, engine='openpyxl') 
    
    for stack_nr in range(0,num_stacks):        
        for column_nr in range(0,num_param):
            column_index = column_indices[column_nr]
            header = column_headers[column_nr]
            data_param = pd.DataFrame({header: data[0][stack_nr][:,column_nr]})
            data_param.to_excel(writer,sheet_name='Sheet'+str(stack_nr),startcol=column_index,index=False)
            
    writer.save()
    
#%% Write array of data to .xlsx file with headers 
    
def write_xlsx_array(file_path,data_array,column_headers,column_indices):
    
    column_indices = range(0,len(column_headers))
    file_name = change_extension(file_path,'.xlsx')
    
    data=pd.DataFrame(data_array)
    num_param = np.shape(data)[1]
    
    try: 
        book = load_workbook(file_name)
        writer = pd.ExcelWriter(file_name, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
    except FileNotFoundError:
        print('No .xlsx file found')        
        writer = pd.ExcelWriter(file_name, engine='openpyxl') 
           
    for column_nr in range(0,num_param):
        column_index = column_indices[column_nr]
        header = column_headers[column_nr]
        data_param = pd.DataFrame({header: data[column_nr]})
        data_param.to_excel(writer,sheet_name='Sheet'+str(0),startcol=column_index,index=False)
            
    writer.save()
    
    return data